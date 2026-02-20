import pandas as pd  # Data manipulation
import numpy as np  # Numerical operations
import matplotlib.pyplot as plt  # Plotting
import joblib  # Model saving
from sklearn.model_selection import GridSearchCV, cross_val_score  # Hyperparameter tuning
from sklearn.preprocessing import LabelEncoder, StandardScaler  # Preprocessing
from sklearn.decomposition import PCA  # Dimensionality reduction
from sklearn.linear_model import LogisticRegression  # Model: LR
from sklearn.tree import DecisionTreeClassifier  # Model: DT
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier  # Model: RF, GB
from sklearn.metrics import accuracy_score, roc_auc_score, log_loss, confusion_matrix, classification_report, roc_curve, f1_score  # Metrics
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ---------------- CONFIGURATION ----------------
TRAIN_URL = "https://raw.githubusercontent.com/ullas9525/Cancer_Cell_Prediction/main/AI_PES%20-%20Inferno%20SplittingofData/AI_PES%20-%20Inferno%20Training_data.xlsx"
TEST_URL = "https://raw.githubusercontent.com/ullas9525/Cancer_Cell_Prediction/main/AI_PES%20-%20Inferno%20SplittingofData/AI_PES%20-%20Inferno%20Testing_data.xlsx"
OUTPUT_DIR = "AI_PES - Inferno OptimizedCode"

# ----------------- 1. LOAD DATA -----------------
print("Loading data from GitHub...")
df_train = pd.read_excel(TRAIN_URL)  # Load train data
df_test = pd.read_excel(TEST_URL)  # Load test data

# ----------------- 2. DATA PREPROCESSING -----------------
leakage_cols = ['Stages', 'Tumor Size', 'Tumor_Age_Ratio', 'Cancer_Type_Malignant']  # Columns to drop
print(f"Dropping leakage columns: {leakage_cols}")

df_train = df_train.drop(columns=[c for c in leakage_cols if c in df_train.columns])  # Drop from train
df_test = df_test.drop(columns=[c for c in leakage_cols if c in df_test.columns])  # Drop from test

# Encode Target
le = LabelEncoder()
df_train['Diagnosis_Status'] = le.fit_transform(df_train['Diagnosis_Status'])  # Encode target
df_test['Diagnosis_Status'] = le.transform(df_test['Diagnosis_Status'])  # Encode target (consistent)

# Separate Features and Target
X_train = df_train.drop('Diagnosis_Status', axis=1)
y_train = df_train['Diagnosis_Status']
X_test = df_test.drop('Diagnosis_Status', axis=1)
y_test = df_test['Diagnosis_Status']

# One-Hot Encoding for Categorical Features
X_train = pd.get_dummies(X_train, drop_first=True)
X_test = pd.get_dummies(X_test, drop_first=True)

# Align Columns (Ensure Train/Test match)
X_train, X_test = X_train.align(X_test, join='left', axis=1, fill_value=0)

# Scaling
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)  # Fit & Transform Train
X_test_scaled = scaler.transform(X_test)  # Transform Test

# ----------------- 3. DIMENSIONALITY REDUCTION (PCA) -----------------
print("Applying PCA...")
pca = PCA(n_components=0.95)  # Keep 95% variance
X_train_pca = pca.fit_transform(X_train_scaled)
X_test_pca = pca.transform(X_test_scaled)
print(f"PCA Reduced dimensions from {X_train_scaled.shape[1]} to {X_train_pca.shape[1]}")

# ----------------- 4. MODEL TRAINING & OPTIMIZATION -----------------
models = {
    "Logistic Regression": {
        "model": LogisticRegression(random_state=42, max_iter=2000),
        "params": {"C": [0.1, 1, 10]}
    },
    "Decision Tree": {
        "model": DecisionTreeClassifier(random_state=42),
        "params": {"max_depth": [3, 5, 7, 10], "min_samples_split": [2, 5]}
    },
    "Random Forest": {
        "model": RandomForestClassifier(random_state=42),
        "params": {"n_estimators": [50, 100, 200], "max_depth": [5, 10, None]}
    },
    "Gradient Boosting": {
        "model": GradientBoostingClassifier(random_state=42),
        "params": {"n_estimators": [50, 100], "learning_rate": [0.01, 0.1, 0.2]}
    }
}

best_overall_score = 0
best_overall_acc = 0
best_overall_model_name = ""
best_overall_model = None
results = []

plt.figure(figsize=(10, 8))  # Setup ROC Plot

for name, config in models.items():
    print(f"\nTraining {name} with GridSearchCV...")
    
    # Grid Search
    grid = GridSearchCV(config["model"], config["params"], cv=5, scoring='accuracy', n_jobs=-1)
    grid.fit(X_train_pca, y_train)
    
    best_model = grid.best_estimator_  # Get best model from grid
    
    # Predictions
    y_pred = best_model.predict(X_test_pca)
    y_prob = best_model.predict_proba(X_test_pca)[:, 1]
    
    # Metrics
    acc = accuracy_score(y_test, y_pred)
    roc = roc_auc_score(y_test, y_prob)
    ll = log_loss(y_test, y_prob)
    f1 = f1_score(y_test, y_pred)
    cv_acc = grid.best_score_  # Best CV score from training
    
    print(f"  Best Params: {grid.best_params_}")
    print(f"  Accuracy: {acc*100:.2f}%")
    print(f"  CV Accuracy: {cv_acc*100:.2f}%")
    print(f"  ROC-AUC: {roc:.4f}")
    print(f"  F1 Score: {f1:.4f}")
    print(f"  Log Loss: {ll:.4f}")
    print("  Confusion Matrix:\n", confusion_matrix(y_test, y_pred))
    print("  Confusion Matrix:\n", confusion_matrix(y_test, y_pred))

    # Append Results
    results.append({
        "Model": name,
        "Accuracy": acc,
        "ROC-AUC": roc,
        "F1 Score": f1,
        "Log Loss": ll
    })
    
    # Store results
    if roc > best_overall_score:  # Selection criteria: ROC-AUC
        best_overall_score = roc
        best_overall_acc = acc
        best_overall_model_name = name
        best_overall_model = best_model

    # Plot ROC
    fpr, tpr, _ = roc_curve(y_test, y_prob)
    plt.plot(fpr, tpr, label=f"{name} (AUC = {roc:.2f})")

# ----------------- 5. FINALIZATION -----------------
# Save ROC Plot
plt.plot([0, 1], [0, 1], 'k--')
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('ROC Curve Comparison')
plt.legend()
roc_path = f"{OUTPUT_DIR}/AI_PES - Inferno ROC_Curve.png"
plt.savefig(roc_path)
print(f"\nROC Curve saved to: {roc_path}")

# Save Best Model
model_path = f"{OUTPUT_DIR}/AI_PES - Inferno Best_Model.pkl"
joblib.dump(best_overall_model, model_path)
print(f"Best Model ({best_overall_model_name}) saved to: {model_path}")

# Save Predictions
# ----------------- 6. VISUALIZATION (CHARTS) -----------------
print("Generating Dashboard (Feature Imp, Distribution, Metrics)...")

# 1. Feature Importance (Helper RF on Original Features)
print("Calculating Feature Importance...")
rf_feat = RandomForestClassifier(random_state=42)
rf_feat.fit(X_train, y_train)
feat_imp_df = pd.DataFrame({
    'Feature': X_train.columns,
    'Importance': rf_feat.feature_importances_
}).sort_values(by='Importance', ascending=False).head(10)  # Top 10 Features
feat_imp_df['Importance'] = feat_imp_df['Importance'].round(2)  # Round to 2 decimals (shorter labels)

# 2. Prediction Confidence Histogram
X_test_final = X_test.copy()
X_test_final['Actual_Diagnosis'] = y_test
X_test_final['Predicted_Diagnosis'] = best_overall_model.predict(X_test_pca)
X_test_final['Prediction_Probability'] = best_overall_model.predict_proba(X_test_pca)[:, 1]

# Calculate Confidence: High if probability is near 0 or 1
X_test_final['Model_Confidence'] = X_test_final['Prediction_Probability'].apply(lambda x: max(x, 1-x))

# Create bins for Confidence (50% to 100%)
bins = [0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
labels = ['50-60%', '60-70%', '70-80%', '80-90%', '90-100%']
X_test_final['Conf_Bin'] = pd.cut(X_test_final['Model_Confidence'], bins=bins, labels=labels, include_lowest=True)

hist_df = X_test_final['Conf_Bin'].value_counts().sort_index().reset_index()
hist_df.columns = ['Confidence Level', 'Count']

# Output Path
pred_path = f"{OUTPUT_DIR}/AI_PES - Inferno FinalPredictionChart.xlsx"

# Create Metrics DataFrame
metrics_df = pd.DataFrame(results).round(2)  # Round all metrics to 2 decimals

# Create Single Sheet layout and Write Tables
with pd.ExcelWriter(pred_path, engine='openpyxl') as writer:
    # A. Predictions (Left Side) - Row 0
    X_test_final.to_excel(writer, sheet_name='Final Analysis', index=False, startrow=0)
    
    # Calculate offset for Right Side components
    offset = len(X_test_final.columns) + 2
    
    # B. Metrics Table (Right Side) - Row 0
    metrics_df.to_excel(writer, sheet_name='Final Analysis', index=False, startrow=0, startcol=offset)
    
    # C. Feature Importance Table (Right Side) - Row 8 (below Metrics)
    feat_imp_df.to_excel(writer, sheet_name='Final Analysis', index=False, startrow=8, startcol=offset)
    
    # D. Histogram Table (Right Side) - Row 22 (below Features)
    hist_df.to_excel(writer, sheet_name='Final Analysis', index=False, startrow=22, startcol=offset)

# Open Workbook to add Charts
wb = load_workbook(pred_path)
ws = wb['Final Analysis']

# Helper to get cell address
def get_cell(col_idx, row_idx):
    return f"{get_column_letter(col_idx)}{row_idx}"

# Common start column for right-side components
start_col = offset + 1  # 1-based index

# --- Chart 1: Metrics Summary (Bar Chart) ---
chart1 = BarChart()
chart1.type = "col"  # Vertical Bars
chart1.title = "Key Performance Metrics"
chart1.y_axis.title = 'Score'
chart1.x_axis.title = 'Model'
chart1.x_axis.tickLblPos = "low" # Force Model Names to appear at bottom
chart1.x_axis.delete = False # Ensure Axis is visible
chart1.x_axis.majorTickMark = "out" # Ensure tick marks are visible
chart1.height = 15
chart1.width = 25

# Define Data and Category references for the chart
categories1 = Reference(ws, min_col=start_col, min_row=2, max_row=len(models)+1)
# Data: Accuracy, ROC, F1 (Col 2, 3, 4 relative to start_col)
data1 = Reference(ws, min_col=start_col+1, min_row=1, max_col=start_col+3, max_row=len(models)+1)

chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(categories1)
# Removed Data Labels for Metrics Chart to prevent collision (Bars are grouped)
ws.add_chart(chart1, get_cell(start_col + 7, 1))  # Place to the right of tables

# --- Chart 2: Feature Importance (Horizontal Bar Chart) ---
chart2 = BarChart()
chart2.type = "bar"  # Horizontal Bars to avoid label overlap
chart2.title = "Top 10 Feature Importance"
chart2.y_axis.title = 'Importance'
chart2.x_axis.title = 'Feature'
chart2.height = 15
chart2.width = 25

# Define Data and Category references for Feature Importance
categories2 = Reference(ws, min_col=start_col, min_row=10, max_row=10+len(feat_imp_df)-1)
data2 = Reference(ws, min_col=start_col+1, min_row=9, max_col=start_col+1, max_row=10+len(feat_imp_df)-1)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(categories2)
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showVal = True
chart2.dataLabels.showCatName = False
chart2.dataLabels.showSerName = False
chart2.dataLabels.position = 'outEnd'
chart2.varyColors = False  # Single color to ensure Axis Labels appear
chart2.legend = None       # Remove legend, force labels on axis
ws.add_chart(chart2, get_cell(start_col + 7, 40)) # PLACEMENT: Row 40 (More space below Chart 1)

# --- Chart 3: Prediction Confidence Histogram (Bar Chart) ---
chart3 = BarChart()
chart3.title = "Model Confidence (How sure is it?)"
chart3.y_axis.title = 'Count of Patients'
chart3.x_axis.title = 'Confidence %'
chart3.height = 15
chart3.width = 25

# Define Data and Category references for Histogram
categories3 = Reference(ws, min_col=start_col, min_row=24, max_row=24+len(hist_df)-1)
data3 = Reference(ws, min_col=start_col+1, min_row=23, max_col=start_col+1, max_row=24+len(hist_df)-1)

chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(categories3)
chart3.varyColors = False  # Single color for clearer axis
chart3.legend = None  # No legend needed for single series
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showVal = True
chart3.dataLabels.showCatName = False
chart3.dataLabels.showSerName = False
chart3.dataLabels.position = 'outEnd' # Push labels outside bar
ws.add_chart(chart3, get_cell(start_col + 7, 80))

wb.save(pred_path)
# Final Summary
print(f"\n>>> BEST MODEL: {best_overall_model_name} with Accuracy: {best_overall_acc*100:.2f}% <<<")